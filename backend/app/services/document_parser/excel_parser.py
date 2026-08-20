from __future__ import annotations

from pathlib import Path

import openpyxl
import xlrd

from app.core.config import settings

from .base import BaseDocumentParser, DocumentParseError, ParsedDocumentResult
from .normalizer import TextNormalizer
from .utils import ensure_safe_zip, ensure_text_limit, to_json_value


class ExcelParser(BaseDocumentParser):
    name = "openpyxl/xlrd"
    extensions = {".xlsx", ".xls"}

    def parse(self, file_path: Path) -> ParsedDocumentResult:
        if file_path.suffix.lower() == ".xls":
            sheets = self._parse_xls(file_path)
        else:
            sheets = self._parse_xlsx(file_path)
        plain_text = self._to_plain_text(sheets)
        return ParsedDocumentResult(
            parser_name=self.name,
            parser_version=self.version,
            plain_text=ensure_text_limit(plain_text),
            sheets=sheets,
            metadata={"sheet_count": len(sheets)},
        )

    def _check_limits(self, sheet_count: int, row_count: int, column_count: int, total_cells: int) -> None:
        if sheet_count > settings.max_excel_sheets:
            raise DocumentParseError(f"Excel Sheet 数超过 {settings.max_excel_sheets} 个限制")
        if row_count > settings.max_excel_rows:
            raise DocumentParseError(f"Excel 行数超过 {settings.max_excel_rows} 行限制")
        if column_count > settings.max_excel_columns:
            raise DocumentParseError(f"Excel 列数超过 {settings.max_excel_columns} 列限制")
        if total_cells > settings.max_excel_cells:
            raise DocumentParseError(f"Excel 单元格总数超过 {settings.max_excel_cells} 个限制")

    def _parse_xlsx(self, file_path: Path) -> list[dict]:
        ensure_safe_zip(file_path)
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
        except (OSError, ValueError, KeyError) as exc:
            raise DocumentParseError("XLSX 文件无效或已损坏") from exc
        try:
            total_cells = 0
            sheets = []
            if len(workbook.worksheets) > settings.max_excel_sheets:
                raise DocumentParseError(f"Excel Sheet 数超过 {settings.max_excel_sheets} 个限制")
            for worksheet in workbook.worksheets:
                rows_count, columns_count = worksheet.max_row, worksheet.max_column
                total_cells += rows_count * columns_count
                self._check_limits(len(workbook.worksheets), rows_count, columns_count, total_cells)
                rows = [
                    [to_json_value(cell) for cell in row]
                    for row in worksheet.iter_rows(values_only=True)
                ]
                sheets.append(
                    {
                        "name": worksheet.title,
                        "rows": rows,
                        "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                    }
                )
            return sheets
        finally:
            workbook.close()

    def _parse_xls(self, file_path: Path) -> list[dict]:
        try:
            workbook = xlrd.open_workbook(file_path, on_demand=True)
        except xlrd.XLRDError as exc:
            raise DocumentParseError("XLS 文件无效或已损坏") from exc
        total_cells = 0
        sheet_count = workbook.nsheets
        if sheet_count > settings.max_excel_sheets:
            raise DocumentParseError(f"Excel Sheet 数超过 {settings.max_excel_sheets} 个限制")
        sheets = []
        for index in range(sheet_count):
            worksheet = workbook.sheet_by_index(index)
            total_cells += worksheet.nrows * worksheet.ncols
            self._check_limits(sheet_count, worksheet.nrows, worksheet.ncols, total_cells)
            rows = [
                [to_json_value(worksheet.cell_value(row, column)) for column in range(worksheet.ncols)]
                for row in range(worksheet.nrows)
            ]
            sheets.append({"name": worksheet.name, "rows": rows, "merged_ranges": []})
        workbook.release_resources()
        return sheets

    @staticmethod
    def _to_plain_text(sheets: list[dict]) -> str:
        text_parts = []
        for sheet in sheets:
            text_parts.append(f"--- Sheet: {sheet['name']} ---")
            text_parts.extend(" | ".join("" if cell is None else str(cell) for cell in row) for row in sheet["rows"])
        return TextNormalizer.normalize("\n".join(text_parts))
